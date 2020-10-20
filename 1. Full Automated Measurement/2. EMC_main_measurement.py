# This script is basically resposible for the measurement.
# It reads the excel sheet created by the previous code(EMC_input_values_old.py), imports the variables and starts the calculation.
# The Spectrum_analyzer.py file contains all the functions that are responsible for the control of the spectrum analyzer.
# The my_instrument variable is made equal to the class that contains the definitions and these definitions are called from this code as well.
# All the functions regarding the spectrum analyzer control starts with "my_instrument." in the code.

import numpy as np
import matplotlib.pyplot as plt
import math
import openpyxl
import serial
import sys
from PIL import Image
import tkinter.ttk as ttk
import tkinter as tk
import tkinter.messagebox
import os
import time
import struct
from Spectrum_analyzer_control import Agilent_N9320B
from Spectrum_analyzer_control import Rigol_DSA800

# If you are using the Rigol analyzer, you should comment out the next line(Agilent_N9320B()) and delete the hashtag sign befor the "my_instrument = Rigol_DSA800()" line.
my_instrument = Agilent_N9320B()
#my_instrument = Rigol_DSA800()

number_of_points_on_the_analyzer = my_instrument.number_of_data_per_measurement()

def round_half_up(n, decimals=0):    # There is no function in python which always rounds the .5 values upwards, therefore I call a definition for this purpuse.
    multiplier = 10 ** decimals
    return math.floor(n*multiplier + 0.5) / multiplier

input_values = openpyxl.load_workbook('Input_Values.xlsx') # to read the excel sheet, which contains the input values.
sheet = input_values.active

#--------------------------------------------------------------------------------------------------------------------------
# These are the input values which are called from the excel sheet. These variables are used during the measurement.
name_of_the_measurement = str(sheet.cell(1, 2).value)
A = float(sheet.cell(4, 2).value)                           # Size of the PCB.
B = float(sheet.cell(5, 2).value)
directory = str(sheet.cell(7, 2).value)            # Indicate the location of PCB's picture which is under test. If there is no picture, do not type anything. The location has to go between the quotation marks. ==> Examle: " "
ID_code_of_spectrum_analyzer = str(sheet.cell(8, 2).value)
Serial_port = str(sheet.cell(9, 2).value)                   # Arduino ==> Tools ==> Port. Here you can check to which serial port arduino is connected and that must to be typed in the quatation mark.
picture_of_the_PCB = str(sheet.cell(10, 2).value)            # The path of the picture
start_freq_level = float(sheet.cell(11, 2).value)            # Hz   Here you can set the desired frequency range
stop_freq_level = float(sheet.cell(12, 2).value)             # Hz
resolution_BW = float(sheet.cell(13, 2).value)               # Hz Here you can set the resolution bandwidth of the spectrum analyzer. 1 MHz gives quick sweep.
step_mm = float(sheet.cell(14, 2).value)                    # The step size that the robot moves between two measurement points.
measurements_in_one_point = int(sheet.cell(15, 2).value)    # The number of measurements in one point.
type_of_the_analysis = str(sheet.cell(16, 2).value)         # maximum, average, single
time_delay_after_measurements = float(sheet.cell(17, 2).value)     # in ms
safety_run = int(sheet.cell(18, 2).value)
#---------------------------------------------------------------------------------------------------------------------------

input_values.save(directory + '/' + name_of_the_measurement + '_input_values.xlsx')
input_values.close()

x = int(round_half_up(A/step_mm) + 1)       # This x value gives the number of cubes on the color map on the horizontal axis
y = int(round_half_up(B/step_mm) + 1)       # This y value gives the number of cubes on the color map on the vertical axis

number_of_measurement_points = int(x * y)   # When the robot stops (after each step) python reads the values from the spectrum analyzer. This gives the number of measurements.

counter = 0  # Variable to help processing the signal from Arduino. It counts the number of measurements.

# These variables are only used to calculate the time duration of the measurement.
measurement_time_in_ms = int((number_of_measurement_points * time_delay_after_measurements * measurements_in_one_point) + (number_of_measurement_points * step_mm * 18.42))
hour    = int(measurement_time_in_ms/(60*60*10**3))
minute  = int(((measurement_time_in_ms/(60*60*10**3))-hour)*60)
second  = int(((((measurement_time_in_ms/(60*60*10**3))-hour)*60)-minute)*60)

ambient_radiation = np.load(directory + '/' + name_of_the_measurement + '_ambient_radiation.npy')
dummy_measurement = np.full(number_of_points_on_the_analyzer, np.min(ambient_radiation)) # The missing measurement points will be filled with the least value of the ambient radiation.

indices_to_be_deleted = np.load(directory + '/' + name_of_the_measurement + '_indices_to_be_deleted.npy') # Loading of the numpy array which contains the number incices to be deleted.

there_is_a_picture = False
if os.path.exists(picture_of_the_PCB) is True:
    there_is_a_picture = True     # if you input the location of a picture it means there will be a background behind the colormap. there_is_a_picture boolean is used to distinguish between displaying a background or no.

my_instrument.connect_device(ID_code_of_spectrum_analyzer)
my_instrument.start_frequency(start_freq_level) # To set the desired frequency range.
my_instrument.stop_frequency(stop_freq_level)
my_instrument.bandwidth(resolution_BW)  # To set the resolution bandwidth on the spectrum analyzer.
my_instrument.auto_sweeptime()          # To set the sweep time of the spectrum analyzer automatically.

frequency = np.linspace(float(start_freq_level), float(stop_freq_level), int(number_of_points_on_the_analyzer))  # This gives the 461 different frequency values which are read from the analyzer. (Agilent analyzer)

#---------------------------------------------------------------------------------------------------------------------------------------
# During the measurement a loading bar is displayed to show the progress of the measurement.
def progress(currentValue):
    progressbar["value"]=currentValue

my_window = tk.Tk()
my_window.title("Automatic measurement")
my_window.geometry("1000x100+200+200")      # Size of the window of the loading bar.

tk.Label(text="Please wait, the measurement is in progress:").pack(side=tk.TOP)
tk.Label(text="The total duration of the measurement is: %d hour %s min %d sec" %(hour, minute, second)).pack(side=tk.BOTTOM)

progressbar=ttk.Progressbar(my_window,orient=tk.HORIZONTAL,length=700,mode='determinate')
progressbar.pack(fill='x', padx=10, pady=20)

currentValue=0
progressbar["value"]=currentValue
progressbar["maximum"]=100

#------------------------------------------------------------------------------------------------------------------------------------
# This part here is responsible for sending the necessary input values to Arduino through the Serial port.
# There are five values, which have to be given to Arduino in order to achieve the proper robot control.
# I am using the serial port communication between the two program.
# The variables to be sent: A, B, step_mm, time_delay_after_measurements, measurements_in_one_point.
# Through the serial port only integers that are between 0-255 can be sent, therefor I need to transform these values in order to make them sendable.
# These values then transformed back to their original shape in Arduino.

A1 = int(str(int(A))[:2])
try:
    A2 = int(str(int(A))[2])
except:
    sys.exc_info()[0]
    A2 = 10

B1 = int(str(int(B))[:2])
try:
    B2 = int(str(int(B))[2])
except:
    sys.exc_info()[0]
    B2 = 10

time1 = int(str(int(time_delay_after_measurements))[:2])
try:
    time2 = int(str(int(time_delay_after_measurements))[2])
except:
    sys.exc_info()[0]
    time2 = 10
try:
    time3 = int(str(int(time_delay_after_measurements))[3])
except:
    sys.exc_info()[0]
    time3 = 10
try:
    time4 = int(str(int(time_delay_after_measurements))[4])
except:
    sys.exc_info()[0]
    time4 = 10

ser = serial.Serial(Serial_port, 9600)# Connecting to the Arduino through the serial monitor.
ser.close()
ser.open()
time.sleep(2)

ser.write(struct.pack('>BBBBBBBBBBB',A1,A2,B1,B2,int(step_mm),time1,time2,time3,time4,int(measurements_in_one_point),safety_run)) # Sending the input values to Arduino.
#---------------------------------------------------------------------------------------------------------------

while True:
    signal = ser.read_all()       # Python starts reading the serial monitor, it makes it possible the send signals from Arduino to python.
    if signal == b'9':  # Arduino sends the number 9 if the emergency button is pressed.
        my_window.destroy()
        my_instrument.close_device()
        tkinter.messagebox.showinfo("Incomplete measurement","The test ended because the Emergency button was pressed.")
        exit()

    if signal == b'8':  # Arduino sends the number 8 if any of the limit switches is pressed.
        my_window.destroy()
        my_instrument.close_device()
        tkinter.messagebox.showinfo("Incomplete measurement", "The measurement ended because one of the limit switches was pressed.\n"
              "You might have given "
              "a too big value for the size of the PCB or did not place the robot to the correct position at the beginning of the test"
              " therefor the robot reached the edge of its frame. ")
        exit()

    if signal == b'6':   # In Arduino it is programmed that after each step it sends a signal (6) to the serial monitor. When python reads this signal, it reads the values from the spectrum analyzer.

        if counter < 1: # At the first measurement I create an array that I call amplitude.
            amplitude = my_instrument.read_the_measurement()
            counter += 1

        else:
            amplitude = np.vstack((amplitude, my_instrument.read_the_measurement())) # Each measurement is written in a seperate row of this amplitude array by stacking the arrays in a sequence vertically.
            counter += 1

            currentValue = (counter/(number_of_measurement_points*measurements_in_one_point))*100 # Updating of the process bar during the measurement in respect to the value of counter.
            progressbar.after(1, progress(currentValue))
            progressbar.update()

        # When the measurement is done, python starts to analyse the measured data.
    if signal == b'3':  # Arduino sends the signal 3 when the measurement is done

        if counter < number_of_measurement_points * measurements_in_one_point:  # If there is missing measurement, the least value of the amplitude array is used to fill the missing points.
            rows_to_fill = (number_of_measurement_points * measurements_in_one_point) - counter
            for i in range(rows_to_fill):
                amplitude = np.vstack((amplitude, dummy_measurement))

        my_window.destroy()
        np.save(directory + '/' + name_of_the_measurement + '_measurement_data.npy', amplitude)
        if os.path.exists(directory + '/' + name_of_the_measurement + '_measurement_data.npy') is True:  # the calculation does not carry on until the file is not saved properly. I need to use this because if the file is big, the saving process might take for a while.
            for k in range (np.size(amplitude,0)):               # all those points where the ambient radiation occure will be overwritten to the smallest value of the amplitude array.
                for i in range(np.size(indices_to_be_deleted)):
                    amplitude[k, indices_to_be_deleted[i][0]] = np.amin(amplitude[0])

            # Progressing the values in respect to the type of the data analysis. (Single, Maximum, Average)
            #---------------------------------------------------------------------------------------------------------------------------------
            if type_of_the_analysis == 'Single':
                dB = amplitude

            elif type_of_the_analysis == 'Maximum':
                dB = np.array([])
                for k in range(number_of_points_on_the_analyzer):
                    dB = np.append(dB, np.amax(amplitude[0:measurements_in_one_point, k])).astype(np.int8)

                for i in range(1, number_of_measurement_points):
                    dummy_array = np.array([])
                    for k in range(number_of_points_on_the_analyzer):
                        dummy_array = np.append(dummy_array, np.amax(amplitude[i * measurements_in_one_point: measurements_in_one_point + i * measurements_in_one_point,k])).astype(np.int8)
                    dB = np.vstack((dB, dummy_array))  # this is also a numpy.ndarray

            elif type_of_the_analysis == 'Average':
                dB = np.array([])
                for k in range(number_of_points_on_the_analyzer):
                    dB = np.append(dB, np.mean(amplitude[0:measurements_in_one_point, k])).astype(np.int8)

                for i in range(1, number_of_measurement_points):
                    dummy_array = np.array([])
                    for k in range(number_of_points_on_the_analyzer):
                        dummy_array = np.append(dummy_array, np.mean(amplitude[i * measurements_in_one_point: measurements_in_one_point + i * measurements_in_one_point,k])).astype(np.int8)
                    dB = np.vstack((dB, dummy_array)) # this is also a numpy.ndarray
                #-----------------------------------------------------------------------------------------------------------------------------------
            np.save(directory + '/' + name_of_the_measurement + '_displayed_data.npy', dB)
            if os.path.exists(directory + '/' + name_of_the_measurement + '_displayed_data.npy') is True: # wait until the file is saved.
                cube = []  # I place the maximum dB value of each measurment point in one array. These values are gonna be displayed on the colormap.
                for k in range(number_of_measurement_points):
                    cube.append(dB[k].max())

                array = np.array(cube).reshape(y,x)  # With this line I reshape the one row cube array which contains the max. dB values to an array which contains x number of rows and y number of columns.

                for k in range(0,y):  # In order to display the measurement points according to the order how they come from the spectrum analyzer, every second row has to be flipped.
                    if k % 2 != 0:  # With this the snake shaped order of the measuring points is obtained.
                        array[k] = np.flip(array[k])

#----------------------------------------------------------------------------------------------------------------------------------------------------------
                def onclick(event):
                    p = int(round((event.xdata + (step_mm / 2)) * (x / (colormap_A+step_mm)) + 0.5))  # This part is to plot the frequency-dB diagrams when clicking on a specific point of the colormap.
                    q = int(round((event.ydata + (step_mm / 2)) * (y / (colormap_B+step_mm)) + 0.5))

                    if q % 2 != 0:
                        q = -1 + (q - 1) * x
                        r = q + p  # r is the value which is filled in the colormap. Actually this indicates the number of measurement point which belongs to the segment of the colormap where you click.
                    else:
                        q = q * x
                        r = q - p

                    deleted_frequencies = np.array([])  # to dtermine the position of the deleted frequencies. In these points a red dot is displayed on the x-axis
                    for k in range(np.size(indices_to_be_deleted)):
                        deleted_frequencies = np.append(deleted_frequencies, frequency[indices_to_be_deleted[k][0]])

                    y_axis_of_red_dots = np.array([])   # the y-axis of the red dots has to be at the zero origin. The origin of the plot is the smallest value of the measured datas minus 5.
                    for k in range(np.size(indices_to_be_deleted)): # this value is placed in an array as many times as many indices are deleted.
                        y_axis_of_red_dots = np.append(y_axis_of_red_dots, np.amin(dB[5]) - 5)

                    plt.figure(2)
                    plt.ylim((np.amin(dB[5]) - 5),0)  # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
                    plt.plot(frequency, dB[r])  # The frequency axis is always the same but the dB values are changing in terms of where the clicking is on the colormap. This r values gives the number of measurement point from where the dB values have to be called.
                    plt.xlabel('Frequency [Hz]')  # These are the labels and title of the plot which comes after clicking on a point of the colormap.
                    plt.ylabel('Amplitude [dB]')
                    plt.title('Emission spectrum')
                    plt.scatter(deleted_frequencies, y_axis_of_red_dots, marker='o', s=10., c='r')  # This is to display the red dots on the edited plot.
                    plt.show()
#-------------------------------------------------------------------------------------------------------------------------------------

                fig = plt.figure(1)  # This is the first figure after analysing the data. This calls the colormap.
                ax = fig.add_subplot()

                alpha = 1  # If there is no background picture, no need to bedim the colormap. In that case the colormap will not be transparent ==> (alpha = 1).
                if there_is_a_picture is True:
                    board = Image.open(picture_of_the_PCB)
                    ax.imshow(board, alpha=1, extent=[0, A, B, 0])
                    alpha = 0.7  # if there is background picture, the colormap is bedimmed a bit

                ax.xaxis.tick_top()
                ax.xaxis.set_label_position('top')

                colormap_A = ((round_half_up(A / step_mm)) * step_mm)
                colormap_B = ((round_half_up(B / step_mm)) * step_mm)

                plt.imshow(array, cmap='YlOrRd', alpha=alpha, extent=[-step_mm / 2, colormap_A + step_mm / 2, colormap_B + step_mm / 2, -step_mm / 2])  # The maximum dB values are saved in the "array" variable.
                clb = plt.colorbar(mappable=None, cax=None, ax=None)
                clb.set_label('Amplitude [dB]', labelpad=0, y=0.5, rotation=90)
                plt.connect('button_press_event', onclick)  # Here I connect the onclick function to this colormap.

                plt.xlabel("A side of the PCB [mm]")
                plt.ylabel("B side of the PCB [mm]")

                plt.tight_layout()
                wm = plt.get_current_fig_manager()  # To display the figure in maximum size.
                wm.window.state('zoomed')

                if counter < number_of_measurement_points * measurements_in_one_point:  # If the measurement was uncomoleated, a window pops up to inform the user.
                    tkinter.messagebox.showinfo("Incomplete measurement",
                                                "The measurement was not completely successful due to %d missing measurement points.\n\n"
                                                "You can find information about this issue in the Error possibilities chapter of the User's guide." % (rows_to_fill))

                plt.show()
                my_instrument.close_device()

                break
