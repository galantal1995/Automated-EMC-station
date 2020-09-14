# This script collects all the input data from the user. These input values are then used with the main script which is doing the measurement.(EMC_measurement.py)
# A GUI interface is designed for the easier use.
# In this code an ambient radiation measurement is also done where the user gets to select those frequency ranges where the ambient radiation occures.
# Two numpy arrays are saved, one which is the ambient radiation and the other one shows the indices of the array where the measured values are not concidered due to the ambient emission.
# An excel sheet is also generated with all the input values which are then transfered to the main script for the further calculations (EMC_measurement.py)
# The Spectrum_analyzer.py file contains all the functions that are responsible for the control of the spectrum analyzer.
# The "my_instrument variable" is being made equal to the class that contains the definitions and these definitions are called from this code as well.
# All the functions regarding the spectrum analyzer control starts with "my_instrument." in the code.

from tkinter import *
import tkinter as tk
import tkinter.filedialog
import tkinter.messagebox
import time
import sys
import math
from openpyxl import Workbook
import cv2
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image, ImageTk
import serial.tools.list_ports
from Spectrum_analyzer import Agilent_N9320B
from Spectrum_analyzer import Rigol_DSA800

# If you are using the Rigol analyzer, you should comment out the next line(Agilent_N9320B()) and delete the hashtag sign befor the "my_instrument = Rigol_DSA800()" line.
my_instrument = Agilent_N9320B()
#my_instrument = Rigol_DSA800()

number_of_points_on_the_analyzer = my_instrument.number_of_data_per_measurement()

final = Workbook()              # To open an excel sheet where the input values will be saved.
sheet_obj = final.active        # In the next following lines the names of these input values are written in the first column.

sheet_obj.cell(row=1, column=1).value = "Name of the board to be measured: "
sheet_obj.cell(row=2, column=1).value = "Material number of the board: "
sheet_obj.cell(row=3, column=1).value = "Additional information about the board, measurement: "
sheet_obj.cell(row=4, column=1).value = "A side of the PCB [mm]: "
sheet_obj.cell(row=5, column=1).value = "B side of the PCB [mm]: "
sheet_obj.cell(row=6, column=1).value = "Height of the probe above the PCB [mm]: "
sheet_obj.cell(row=7, column=1).value = "Folder where the measurement is saved: "
sheet_obj.cell(row=8, column=1).value = "Identification number of the Spectrum Analyzer: "
sheet_obj.cell(row=9, column=1).value = "Serial Port of the plotter: "
sheet_obj.cell(row=10, column=1).value = "Picture of the PCB board: "
sheet_obj.cell(row=11, column=1).value = "Desired start frequency of the analyzer [Hz]: "
sheet_obj.cell(row=12, column=1).value = "Desired stop frequency of the analyzer [Hz]: "
sheet_obj.cell(row=13, column=1).value = "Resolution bandwidth [Hz]: "
sheet_obj.cell(row=14, column=1).value = "Step size [mm]: "
sheet_obj.cell(row=15, column=1).value = "Number of measurements in one point: "
sheet_obj.cell(row=16, column=1).value = "Type of the analysis in each point: "
sheet_obj.cell(row=17, column=1).value = "Time delay after each measurement: [ms]: "
sheet_obj.cell(row=18, column=1).value = "Safety run (1==No, 2==Yes): "

def round_half_up(n, decimals=0):       # There is no function in python which always rounds the .5 values upwards, therefore I call a definition for this purpuse.
    multiplier = 10 ** decimals
    return math.floor(n*multiplier + 0.5) / multiplier

def connected_board(): # Automaticly defines the usb port to which the MeOrion plotter control board is connected.
    usb_devices = serial.tools.list_ports.comports(include_links=True)
    for i in range(len(usb_devices)):
        if usb_devices[i][1][:16] == "USB-SERIAL CH340":
            return(usb_devices[i][0])  # You can manually realize the connection with entering: return "COM4" (IF YOUR BOARD IS CONNECTED TO COM4!)

# In the next following lines definitions are called which are responsible for the functions of the different buttons.
#--------------------------------------------------------------------------------------------------------------------------------
def path_to_save():  # Print path button. Selection of the folder where the measurement is saved.
    directory = tk.filedialog.askdirectory()
    e7.delete(0, tk.END)
    e7.insert(0, directory)  # It writes the path of the selected directory into the 6th typing field.

def ID_of_analyzer():   # Select button.
    try:
        my_instrument.find_ID()
        e8.delete(0, tk.END)
        e8.insert(0, my_instrument.getID())
        tkinter.messagebox.showinfo("Spectrum Analyzer","The following measuring device is connected to the computer:\n\n" + (str(my_instrument.info_on_device())))
        my_instrument.connect_device(e8.get())
    except:
        sys.exc_info()[0]
        tk.messagebox.showinfo("No device found","Please, check the connection of the spectrum analyzer because there is no detected device.")
        pass

def path_of_picture():   # Print path button. Selection of the picture which will be displayed behind the colormap.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e9.delete(0, tk.END)
    e9.insert(0, path)

# The following part is only responsible for editing the imported picture. It makes it possible to cut and edit the picture and saves the edited version. This edited picure will be displayed.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------
def mouse_crop(event, x, y, flags, param):
    name_of_the_measurement = e1.get()
    Picture = e9.get()
    Folder = e7.get()
    # grab references to the global variables
    oriImage = cv2.imread(str(Picture))
    global x_start, y_start, x_end, y_end, cropping
    # if the left mouse button was DOWN, start RECORDING
    # (x, y) coordinates and indicate that cropping is being
    if event == cv2.EVENT_LBUTTONDOWN:
        x_start, y_start, x_end, y_end = x, y, x, y
        cropping = True
    # Mouse is Moving
    elif event == cv2.EVENT_MOUSEMOVE:
        if cropping == True:
            x_end, y_end = x, y
    # if the left mouse button was released
    elif event == cv2.EVENT_LBUTTONUP:
        # record the ending (x, y) coordinates
        x_end, y_end = x, y
        cropping = False  # cropping is finished
        refPoint = [(x_start, y_start), (x_end, y_end)]
        if len(refPoint) == 2:  # when two points were found
            roi = oriImage[refPoint[0][1]:refPoint[1][1], refPoint[0][0]:refPoint[1][0]]
            cv2.imshow("Cropped image. Please press ESC to save and close the pictures.", roi)
            path_of_the_new_picture = str(Folder) + "/" + str(name_of_the_measurement) + '_picture.png'
            cv2.imwrite(path_of_the_new_picture, roi)
            e9.delete(0, tk.END)
            e9.insert(0, path_of_the_new_picture)

cropping = False
x_start, y_start, x_end, y_end = 0, 0, 0, 0

def editing_the_picture():
    Picture = e9.get()
    image = cv2.imread(str(Picture))
    cv2.namedWindow("image")
    cv2.setMouseCallback("image", mouse_crop)
    while True:
        i = image.copy()
        if not cropping:
            cv2.imshow("image", image)
            if cv2.waitKey(1) == 27:
                break
        elif cropping:
            cv2.rectangle(i, (x_start, y_start), (x_end, y_end), (255, 0, 0), 2)
            cv2.imshow("image", i)
            if cv2.waitKey(1) == 27:
                break
    cv2.destroyAllWindows()
#--------------------------------------------------------------------------------------------------------------------------------------------------


# Ambient radiation measurement
#---------------------------------------------------------------------------------------------------------------------------------------------------
def Start():   # Measurement of the ambient radiation.
    response = tkinter.messagebox.askquestion("Information", "After clicking on the yes button, the ambient readiation measurement will start.\nThe process will take 1 minute. "
                                               "Please wait until you get a command from the computer to proceed.")
    if 'yes' in response:
        my_instrument.start_frequency(e10.get())
        my_instrument.stop_frequency(e11.get())
        my_instrument.bandwidth(e12.get())
        my_instrument.auto_sweeptime()
        my_instrument.maxhold_on()
        time.sleep(60)                  # Here you can set the time duration of the max hold.
        ambient_radiation = my_instrument.read_the_measurement()
        np.save(str(e7.get()) + '/' + str(e1.get()) + '_ambient_radiation.npy', ambient_radiation)  # saving the ambient radiation into the previously given folder.
        frequency = np.linspace(float(e10.get()), float(e11.get()), int(number_of_points_on_the_analyzer)) # these are the corresponding frequency values to the dB vales taken from the analyzer. I create numpy arrays from both of them for the easier processing.
        my_instrument.maxhold_off()
        plt.figure(1)
        plt.connect('button_press_event', onclick_ambient) # connection of the plot with the onclick_ambient function.
        plt.ylim((np.amin(ambient_radiation) - 5),0)  # To set the y axis to be between 0 and the least value of the amplidudes. It makes it easier to visualize.
        plt.plot(frequency, ambient_radiation)  # Plotting the result of the max hold. The same picture appears on the computer as it is on the analyzer.
        plt.xlabel('Frequency [Hz]')
        plt.ylabel('Amplitude [dB]')
        plt.title('Ambient radiation')
        plt.grid()
        plt.show()

# To select the frequency ranges to be deleted you must click on the plot of the ambient radiation.
# What happens after the clicking is realized here in the onclick_ambient function.
def onclick_ambient(event):
    selected_maximum_amplitude = event.ydata    # the maximum amplitude value is given by the position of where the clicking happens. All those points which are greater then this max value are going to be deleted.
    ambient_radiation = np.load(str(e7.get()) + '/' + str(e1.get()) + '_ambient_radiation.npy')
    indices_to_be_deleted = np.argwhere(ambient_radiation > selected_maximum_amplitude)   # it gives those points(indices of array) which are greater than the selected maximum.
    np.save(str(e7.get()) + '/' + str(e1.get()) + '_indices_to_be_deleted.npy', indices_to_be_deleted) # these indices are saved
    response = tkinter.messagebox.askquestion("Attenition!","Please pay attention that the number of ignored points is: %d\n"
                                                         "Would you like to proceed?" %(np.size(indices_to_be_deleted)))
    if 'yes' in response:
        np.save(str(e7.get()) + '/' + str(e1.get()) + '_indices_to_be_deleted.npy', indices_to_be_deleted)  # these indices are saved
        for i in range(np.size(indices_to_be_deleted)):
            ambient_radiation[indices_to_be_deleted[i][0]] = np.amin(ambient_radiation) # I replace those points which are above the maximum with the smallest value of the arra.
        frequency = np.linspace(float(e10.get()), float(e11.get()), int(number_of_points_on_the_analyzer))
        e13.delete(0, tk.END)
        e13.insert(0, "Done!")

        deleted_frequencies = np.array([])      # to dtermine the position of the deleted frequencies. In these points a red dot is displayed on the x-axis
        for k in range(np.size(indices_to_be_deleted)):
            deleted_frequencies = np.append(deleted_frequencies, frequency[indices_to_be_deleted[k][0]])

        y_axis_of_red_dots = np.array([])   # the y-axis of the red dots has to be at the zero origin. The origin of the plot is the smallest value of the measured datas minus 5.
        for k in range(np.size(indices_to_be_deleted)): # this value is placed in an array as many times as many indices are deleted.
            y_axis_of_red_dots = np.append(y_axis_of_red_dots, np.amin(ambient_radiation) - 5)

        plt.figure(2)
        plt.ylim((np.amin(ambient_radiation) - 5),0)  # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
        plt.plot(frequency,ambient_radiation)  # The frequency axis is always the same but the dB values are changing in terms of where the clicking is on the colormap. This r values gives the number of measurement point from where the dB values have to be called.
        plt.xlabel('Frequency [Hz]')  # These are the labels and title of the plot which comes after clicking on a point of the colormap.
        plt.ylabel('Amplitude [dB]')
        plt.title('Ambient radiation (edited)')
        plt.scatter(deleted_frequencies, y_axis_of_red_dots, marker='o', s=10., c='r')  # This is to display the red dots on the edited plot.
        plt.show()
#-------------------------------------------------------------------------------------------------------------------------------------------------------------------

# The calculate button sets the desired frequency range and bandwidth to the spectrum analyzer and reads the sweep time of it.
# It calculates a necessary minimum time delay between two measurements.
def calculate():
    # The value of these variables are given by the user and then collected from the input fields for further calculations. (e7.get ==> read the input field)

    my_instrument.start_frequency(e10.get())
    my_instrument.stop_frequency(e11.get())
    my_instrument.bandwidth(e12.get())
    my_instrument.auto_sweeptime()
    sweep_time = my_instrument.read_sweeptime()
    time_delay = int((sweep_time + 10)) # I add 10 ms to the sweep time

    A = e4.get()
    B = e5.get()
    step = e14.get()
    number_of_measurement_in_one_point = int(e15.get())
    a = float(float(A) + float(step))
    b = float(float(B) + float(step))
    x = int(round_half_up(a / float(step)))  # This x value gives the number of cubes on the color map on the horizontal axis
    y = int(round_half_up(b / float(step)))  # This y value gives the number of cubes on the color map on the vertical axis
    number_of_measurement_points = int(x * y)  # When the robot stops (after each step) python reads the values from the spectrum analyzer. This gives the number of measurement points.
    total_number_of_measurements = number_of_measurement_points * number_of_measurement_in_one_point
    minimum_delay = round_half_up(1.73984214e-15*total_number_of_measurements**3 - 1.5330159e-9*total_number_of_measurements**2 + 6.40960304e-4*total_number_of_measurements+1.00463828e2) # polynomial equation to determine the minimum time delay regarding the number of measurements.

    if time_delay < minimum_delay: # depending on the number of measurements, the computer needs a certain amount of time between two measurements to execute the necessary operations.
        time_delay = minimum_delay + 60

    e17.delete(0, tk.END)
    e17.insert(0, int(time_delay))

    tkinter.messagebox.showinfo("Attention!", "Please, power on the PCB board to be tested.")

def continue_button():
    name_of_the_measurement = e1.get()
    material_number = e2.get()
    additional_info = e3.get()
    A = e4.get()
    B = e5.get()
    height_of_probe = e6.get()
    Picture = e9.get()
    ID = e8.get()
    Serial_port = connected_board()
    Folder = e7.get()
    Start_freq = e10.get()
    Stop_freq = e11.get()
    BW = e12.get()
    step = e14.get()
    number_of_measurement_in_one_point = e15.get()

    a = float(float(A) + float(step))
    b = float(float(B) + float(step))

    x = int(round_half_up(a / float(step)))  # This x value gives the number of cubes on the color map on the horizontal axis
    y = int(round_half_up(b / float(step)))  # This y value gives the number of cubes on the color map on the vertical axis

    number_of_measurement_points = int(x * y)  # When the robot stops (after each step) python reads the values from the spectrum analyzer. This gives the number of measurements.

    if (int(number_of_measurement_in_one_point)*number_of_measurement_points) > 300000:    # Test with 300 thousand measurement points hasnt been tested yet.
        tkinter.messagebox.showinfo("Attention!", "This setup would result in a too big number of measurement points, which hasn't been tested before."
                                                  "\n\nCheck if you entered the size of the PCB correctly and if yes, "
                                                  "please, increase the step size or decrease the number of measurements in one point and click on the continue button again.")

    else:
        if e16.get() == 1:
            type_of_analysis = "Single"
        elif e16.get() == 2:
            type_of_analysis = "Maximum"
        elif e16.get() == 3:
            type_of_analysis = "Average"

        if e15.get() == str(1): # If you only want to do one measurement in one point then the analysis type automaticaly turns to be Single.
            type_of_analysis = "Single"
        if int(number_of_measurement_in_one_point) > 1 and type_of_analysis != "Average":
            type_of_analysis = "Maximum"
        time_delay = e17.get()

        # It copies all the input data from the typing fields and writes them in the excel sheet.
        sheet_obj.cell(row=1, column=2).value = name_of_the_measurement
        sheet_obj.cell(row=2, column=2).value = material_number
        sheet_obj.cell(row=3, column=2).value = additional_info
        sheet_obj.cell(row=4, column=2).value = A
        sheet_obj.cell(row=5, column=2).value = B
        sheet_obj.cell(row=6, column=2).value = height_of_probe
        sheet_obj.cell(row=7, column=2).value = Folder
        sheet_obj.cell(row=8, column=2).value = ID
        sheet_obj.cell(row=9, column=2).value = Serial_port
        sheet_obj.cell(row=10, column=2).value = Picture
        sheet_obj.cell(row=11, column=2).value = Start_freq
        sheet_obj.cell(row=12, column=2).value = Stop_freq
        sheet_obj.cell(row=13, column=2).value = BW
        sheet_obj.cell(row=14, column=2).value = step
        sheet_obj.cell(row=15, column=2).value = number_of_measurement_in_one_point
        sheet_obj.cell(row=16, column=2).value = type_of_analysis
        sheet_obj.cell(row=17, column=2).value = time_delay
        sheet_obj.cell(row=18, column=2).value = e18.get()      # Indicates the existance of the safety run.
        final.save('Input_values.xlsx')
        final.close()


        measurement_time_in_ms = int((number_of_measurement_points * float(time_delay) * float(number_of_measurement_in_one_point)) + (number_of_measurement_points * float(step) * 18.42))
        hour = int(measurement_time_in_ms / (60 * 60 * 10 ** 3))
        minute = int(((measurement_time_in_ms / (60 * 60 * 10 ** 3)) - hour) * 60)
        second = int(((((measurement_time_in_ms / (60 * 60 * 10 ** 3)) - hour) * 60) - minute) * 60)

        if e18.get() == 2:
            response = tk.messagebox.askquestion("Starting the measurement", "After clicking on the Yes button a safety run will start. "
                                                                        "\nDuring this time please keep an eye on the probe and if you anticipate the possibility of an accidental collision, please push the emergency button instantly."
                                                                        "\n\nThe measurement will start right after the safety run and will take %d hours %d minutes %d seconds"
                                                                        "\n\nDo you want to start the process?" %(hour, minute, second))
        elif e18.get() == 1:
            response = tk.messagebox.askquestion("Starting the measurement", "After clicking on the Yes button, the measurement will start.\n"
                                                                            "The total duration of the test is: %d hours %d minutes %d seconds"
                                                                            "\n\nDo you want to start the process?" %(hour, minute, second))

        if 'yes' in response:
            my_window.destroy()  # When you click on the yes button, the input window will close and the EMC_measurement.py script will run.
            plt.close('all')
            import EMC_measurement

def exit_now():  # Quit button
    my_window.quit()


# Here the displayed window is being designed.
#----------------------------------------------------------------------------------------------------------------------------------------------
my_window = tk.Tk()
my_window.title("Automatic measurement system")
my_window.state('zoomed')
font = 2        # If your screen is much smaller just decrease the font size of the letters and the window will be displayable.
tk.Label(my_window, text="1. Name of the board to be measured: ", font=font).grid(row=2, column=1)
tk.Label(my_window, text="2. Material number of the board: ", font=font).grid(row=3, column=1)
tk.Label(my_window, text="3. Additional information about the board, measurement: ", font=2).grid(row=4, column=1)
tk.Label(my_window, text="4. A side of the PCB [mm]: ", font=font).grid(row=5, column=1)
tk.Label(my_window, text="5. B side of the PCB [mm]: ", font=font).grid(row=6, column=1)
tk.Label(my_window, text="6. Height of the probe above the PCB [mm]: ", font=font).grid(row=7, column=1)
tk.Label(my_window, text="7. Select the folder where the measurement shall be saved: ", font=font).grid(row=8, column=1)
tk.Label(my_window, text="8. Identification number of the Spectrum Analyzer: ", font=font).grid(row=9, column=1)
tk.Label(my_window, text="9. Picture of the PCB board: ", font=font).grid(row=10, column=1)
tk.Label(my_window, text="10. Desired start frequency of the analyzer [Hz]: ", font=font).grid(row=11, column=1)
tk.Label(my_window, text="11. Desired stop frequency of the analyzer [Hz]: ", font=font).grid(row=12, column=1)
tk.Label(my_window, text="12. Resolution bandwidth [Hz]: ", font=font).grid(row=13, column=1)
tk.Label(my_window, text="13. Measurement of the ambient radiation: ", font=font).grid(row=14, column=1)
tk.Label(my_window, text="14. Step size [mm]: ", font=font).grid(row=15, column=1)
tk.Label(my_window, text="15. Number of measurements in one point: ", font=font).grid(row=16, column=1)
tk.Label(my_window, text="16. Type of the analysis in each point: ", font=font).grid(row=17, column=1)
tk.Label(my_window, text="17. Time delay after each measurement [ms]: ", font=font).grid(row=18, column=1)
tk.Label(my_window, text="18. Would you like the robot to perform a safety run first? ", font=font).grid(row=19, column=1)
tk.Label(my_window, text="", font=font).grid(row=20, column=1) # Empty rows before quit and continue button.
tk.Label(my_window, text="", font=font).grid(row=21, column=1)

e1 = tk.Entry(my_window)  # Setting of the input fields.
e2 = tk.Entry(my_window)
e3 = tk.Entry(my_window)
e4 = tk.Entry(my_window)
e5 = tk.Entry(my_window)
e6 = tk.Entry(my_window)
e7 = tk.Entry(my_window)
e8 = tk.Entry(my_window)
e10 = tk.Entry(my_window)
e9 = tk.Entry(my_window)
e11 = tk.Entry(my_window)
e12 = tk.Entry(my_window)
e13 = tk.Entry(my_window)
e14 = tk.Entry(my_window)
e15 = tk.Entry(my_window)
e16 = IntVar(my_window)    # Checkbox of the 16th question. It can be either Single or Maximum or Average
e17 = tk.Entry(my_window)
e18 = IntVar(my_window)    # Checkbox of the 18th question. Selection of safety run

e1.grid(row=2, column=2, pady=6)    # Input grid fields
e2.grid(row=3, column=2, pady=6)
e3.grid(row=4, column=2, pady=6)
e4.grid(row=5, column=2, pady=6)
e5.grid(row=6, column=2, pady=6)
e6.grid(row=7, column=2, pady=6)
e7.grid(row=8, column=2, pady=6)
e8.grid(row=9, column=2, pady=6)
e9.grid(row=10, column=2, pady=6)
e10.grid(row=11, column=2, pady=6)
e11.grid(row=12, column=2, pady=6)
e12.grid(row=13, column=2, pady=6)
e13.grid(row=14, column=2, pady=6)
e14.grid(row=15, column=2, pady=6)
e15.grid(row=16, column=2, pady=6)
e17.grid(row=18, column=2, pady=6)

# Settings of the buttons. All the buttons call a separate definition depending on the function it stands for.
tk.Button(my_window, text='Print path', font=font, command=path_to_save, width=8).grid(row=8, column=3, padx=3, pady=3)
tk.Button(my_window, text='Select', font=font, command=ID_of_analyzer, width=8).grid(row=9, column=3, padx=3, pady=3)
tk.Button(my_window, text='Print path', font=font, command=path_of_picture, width=8).grid(row=10, column=3, padx=3, pady=3)
tk.Button(my_window, text='Edit', font=font, command=editing_the_picture, width=8).grid(row=10, column=4, padx=3, pady=3)
tk.Button(my_window, text='Start', font=font, command=Start, width=8).grid(row=14, column=3, padx=3, pady=3)
Radiobutton(my_window, text="Single", font=font, variable=e16, value=1).grid(row=17, column=2, padx=3, pady=3)
Radiobutton(my_window, text="Maximum", font=font, variable=e16, value=2).grid(row=17, column=3, padx=3, pady=3)
Radiobutton(my_window, text="Average", font=font, variable=e16, value=3).grid(row=17, column=4, padx=3, pady=3)
tk.Button(my_window, text='Calculate', font=font, command=calculate, width=8).grid(row=18, column=3, padx=3, pady=3)
Radiobutton(my_window, text="No", font=font, variable=e18, value=1).grid(row=19, column=2, padx=3, pady=3)
Radiobutton(my_window, text="Yes", font=font, variable=e18, value=2).grid(row=19, column=3, padx=3, pady=3)
tk.Button(my_window, text='Continue', font=font, command=continue_button, width=8).grid(row=22, column=4, padx=3, pady=3)
tk.Button(my_window, text='Quit', font=font,command=exit_now, width=8).grid(row=22, column=1, padx=3, pady=3)

# If your screen is big enough, uncomment these two lines and a tesat logo will appear on the starting window.
#tesat = ImageTk.PhotoImage(Image.open("tesat.jpg"))
#tk.Label(my_window, image=tesat).grid(row=1, column=6) # Tesat image on the right side.

my_window.grid_rowconfigure(0, weight=1)   # arranging the the content of the window in the center
my_window.grid_rowconfigure(23, weight=1)
my_window.grid_columnconfigure(0, weight=1)
my_window.grid_columnconfigure(7, weight=1)

my_window.mainloop()