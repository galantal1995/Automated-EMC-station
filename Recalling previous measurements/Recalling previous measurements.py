# This script makes it possible to recall previously executed measurements.
# In order to use this, you must have the files which were saved from the measurement.
# This one also includes a gui interface but i does not call any other scripts.
# The calculation is done when you click on the Continue button. (Main calculation is within the def Calculation())

from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
import tkinter.messagebox
import numpy as np
import matplotlib.pyplot as plt
import math
import openpyxl
from PIL import Image, ImageTk
import os
import sys
from Spectrum_analyzer import Agilent_N9320B
from Spectrum_analyzer import Rigol_DSA800

# If you are using the Rigol analyzer, you should comment out the next line(Agilent_N9320B()) and delete the hashtag sign befor the "my_instrument = Rigol_DSA800()" line.
my_instrument = Agilent_N9320B()
#my_instrument = Rigol_DSA800()

number_of_points_on_the_analyzer = my_instrument.number_of_data_per_measurement()

def round_half_up(n, decimals=0):    # There is no function in python which always rounds the .5 values upwards, therefore I call a definition for this purpuse.
    multiplier = 10 ** decimals
    return math.floor(n*multiplier + 0.5) / multiplier

def print_path_input_values(): # Print path button. Selecting the excel file which contains the input values.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e1.delete(0, tk.END)
    e1.insert(0, path)

def print_path_measurement_data(): # Print path button. Selecting the numpy array file which contains the measurement values.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e2.delete(0, tk.END)
    e2.insert(0, path)

def print_path_displayed_data():
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e2.delete(0, tk.END)
    e2.insert(0, path)

def print_path_ambient_radiation(): # Print path button. Selecting the numpy array file which contains the ambient radiation.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e3.delete(0, tk.END)
    e3.insert(0, path)

def print_path_indices_to_be_deleted(): # Print path button. Selecting the numpy array file which contains the indices of array which are ignored.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e4.delete(0, tk.END)
    e4.insert(0, path)

def print_path_picture(): # Print path button. Importing the picture of the PCB.
    path = tk.filedialog.askopenfilename(parent=my_window, initialdir='C:/Tutorial',title='Choose file',filetypes=[('All Files', '.*'),('png images', '.png'),('gif images', '.gif'), ('jpg images', '.jpg')])
    e5.delete(0, tk.END)
    e5.insert(0, path)

def selection_of_ambient_radiation():  # Displaying the ambient radiation which was measured before the measurement started.
    ambient_radiation = np.load(str(e3.get()))
    input_values_path = str(e1.get())
    input_values = openpyxl.load_workbook(input_values_path)     # Importing the input values.
    sheet = input_values.active
    start_freq_level = float(sheet.cell(11, 2).value)
    stop_freq_level = float(sheet.cell(12, 2).value)
    frequency = np.linspace(float(start_freq_level), float(stop_freq_level), int(number_of_points_on_the_analyzer))  # Processing the values from the excel sheet into arrays for the easier analysation.
    plt.connect('button_press_event', onclick_evaluation)
    plt.ylim((np.amin(ambient_radiation) - 5),0)  # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
    plt.plot(frequency, ambient_radiation)  # Plotting the result of the max hold. The same picture appears on the computer as it is on the analyzer.
    plt.xlabel('Frequency [Hz]')
    plt.ylabel('Amplitude [dB]')
    plt.title('Ambient radiation')
    plt.grid()
    plt.show()

def onclick_evaluation(event):
    # To select the frequency ranges to be deleted you must click on the plot of the ambient radiation.
    # What happens after the clicking is realized here in the onclick_evaluation funcion.
    selected_maximum_amplitude = event.ydata    # the maximum amplitude value is given by the position of where the clicking happens. All those points which are greater then this max value are going to be deleted.
    ambient_radiation = np.load(str(e3.get()))
    indices_to_be_deleted = np.argwhere(ambient_radiation > selected_maximum_amplitude)

    response = tkinter.messagebox.askquestion("Attenition!","Please pay attention that the number of ignored points is: %d\n"
                                                            "Would you like to proceed?" %(np.size(indices_to_be_deleted)))
    if 'yes' in response:
        np.save(str(e4.get()), indices_to_be_deleted)
        for i in range(np.size(indices_to_be_deleted)):     # I replace those points which are above the maximum with the smallest value of the arra.
            ambient_radiation[indices_to_be_deleted[i][0]] = np.amin(ambient_radiation)

        input_values_path = str(e1.get())
        input_values = openpyxl.load_workbook(input_values_path)     # Importing the input values.
        sheet = input_values.active
        start_freq_level = float(sheet.cell(11, 2).value)
        stop_freq_level = float(sheet.cell(12, 2).value)
        frequency = np.linspace(float(start_freq_level), float(stop_freq_level), int(number_of_points_on_the_analyzer))  # Processing the values from the excel sheet into arrays for the easier analysation.

        e6.delete(0, tk.END)
        e6.insert(0, "Done!")
        deleted_frequencies = np.array([])      # to dtermine the position of the deleted frequencies. In these points a red dot is displayed on the x-axis
        for k in range(np.size(indices_to_be_deleted)):
            deleted_frequencies = np.append(deleted_frequencies, frequency[indices_to_be_deleted[k][0]])

        y_axis_of_red_dots = np.array([])   # the y-axis of the red dots has to be at the zero origin. The origin of the plot is the smallest value of the measured datas minus 5.
        for k in range(np.size(indices_to_be_deleted)):
            y_axis_of_red_dots = np.append(y_axis_of_red_dots, np.amin(ambient_radiation) - 5)

        plt.figure(2)
        plt.ylim((np.amin(ambient_radiation) - 5),0)  # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
        plt.plot(frequency,ambient_radiation)  # The frequency axis is always the same but the dB values are changing in terms of where the clicking is on the colormap. This r values gives the number of measurement point from where the dB values have to be called.
        plt.xlabel('Frequency [Hz]')  # These are the labels and title of the plot which comes after clicking on a point of the colormap.
        plt.ylabel('Amplitude [dB]')
        plt.title('Ambient radiation (edited')
        plt.scatter(deleted_frequencies, y_axis_of_red_dots, marker='o', s=10., c='r')  # This is to display the red dots on the edited plot.
        plt.show()

def Continue_with_changes():  # This loop gets executed when clicking on the continue button if the user chose to change on the consideration of ambient radiation or on the analyzation type
    plt.close('all')
    input_values_path = str(e1.get())
    measurement_data_path = str(e2.get())
    indices_to_be_deleted_path = str(e4.get())
    picture_of_the_PCB = str(e5.get())
    type_of_the_analysis = int(e7.get())

    #-------------------------------------------------------------------------------------------------
    input_values = openpyxl.load_workbook(input_values_path)     # Importing the input values.
    sheet = input_values.active
    A = float(sheet.cell(4, 2).value)
    B = float(sheet.cell(5, 2).value)
    start_freq_level = float(sheet.cell(11, 2).value)
    stop_freq_level = float(sheet.cell(12, 2).value)
    step_mm = float(sheet.cell(14, 2).value)
    measurements_in_one_point = int(sheet.cell(15, 2).value)
    Quit()
    if measurements_in_one_point == 1:
        type_of_the_analysis = 1

    if measurements_in_one_point > 1 and type_of_the_analysis != 3:
        type_of_the_analysis = 2

    x = int(round_half_up(A / step_mm) + 1)  # This x value gives the number of cubes on the color map on the horizontal axis
    y = int(round_half_up(B / step_mm) + 1)  # This y value gives the number of cubes on the color map on the vertical axis

    number_of_measurement_points = int(x * y)  # When the robot stops (after each step) python reads the values from the spectrum analyzer. This gives the number of measurements.

    indices_to_be_deleted = np.load(indices_to_be_deleted_path) # Loading of the numpy array which contains the number incices to be deleted.

    there_is_a_picture = False

    if os.path.exists(picture_of_the_PCB) is True:
        there_is_a_picture = True  # if you input the location of a picture it means there will be a background behind the colormap. there_is_a_picture boolean is used to distinguish between displaying a background or no.

    frequency = np.linspace(float(start_freq_level), float(stop_freq_level), int(number_of_points_on_the_analyzer))  # Processing the values from the excel sheet into arrays for the easier analysation.

    amplitude = np.load(measurement_data_path)    # to load the measured data

    for k in range(np.size(amplitude,0)):  # all those points where the ambient radiation occure will be overwritten to the smallest value of the amplitude array.
        for i in range(np.size(indices_to_be_deleted)):
            amplitude[k, indices_to_be_deleted[i][0]] = np.amin(amplitude[0])

    # Progressing the values in respect to the type of the data analysis. (Single, Maximum, Average)
    #----------------------------------------------------------------------------------------------------------------
    if type_of_the_analysis == 2 or type_of_the_analysis == 3:  # Processing the data might take longer, therefore I call a loading bar in case of the maximum and average analysis.
        def progress(currentValue):                             # This loading bar shows the state of the process to the user.
            progressbar["value"] = currentValue
        my_window2 = tk.Tk()
        my_window2.title("Loading of the previously executed measurement")
        my_window2.geometry("1000x100+200+200")  # Size of the window of the loading bar.
        tk.Label(text="The measurement data is being processed, please, wait.").pack(side=tk.TOP)
        progressbar = ttk.Progressbar(my_window2, orient=tk.HORIZONTAL, length=700, mode='determinate')
        progressbar.pack(fill='x', padx=10, pady=20)
        currentValue = 0
        progressbar["value"] = currentValue
        progressbar["maximum"] = 100

    if type_of_the_analysis == 1: # It is a single type of analyzes with only one measurement in one point
        dB = amplitude

    elif type_of_the_analysis == 2: # Progressbar is only displayed  when the processing type is maximum or average due to the 1-2 minute data processing time in this loop.
        dB = np.array([])
        for k in range(number_of_points_on_the_analyzer):
            dB = np.append(dB, np.amax(amplitude[0:measurements_in_one_point, k])).astype(np.int8)

        for i in range(1, number_of_measurement_points):
            dummy_array = np.array([])
            for k in range(number_of_points_on_the_analyzer):
                dummy_array = np.append(dummy_array, np.amax(amplitude[i * measurements_in_one_point: measurements_in_one_point + i * measurements_in_one_point,k])).astype(np.int8)
            dB = np.vstack((dB, dummy_array))  # this is also a numpy.ndarray
            currentValue = (i / number_of_measurement_points) * 100  # Updating of the process bar.
            progressbar.after(1, progress(currentValue))
            progressbar.update()

    elif type_of_the_analysis == 3: # Progressbar is only displayed  when the processing type is maximum or average due to the 1-2 minute data processing time in this loop.
        dB = np.array([])
        for k in range(number_of_points_on_the_analyzer):
            dB = np.append(dB, np.mean(amplitude[0:measurements_in_one_point, k])).astype(np.int8)

        for i in range(1, number_of_measurement_points):
            dummy_array = np.array([])
            for k in range(number_of_points_on_the_analyzer):
                dummy_array = np.append(dummy_array, np.mean(amplitude[i * measurements_in_one_point: measurements_in_one_point + i * measurements_in_one_point,k])).astype(np.int8)
            dB = np.vstack((dB, dummy_array))  # this is also a numpy.ndarray
            currentValue = (i / number_of_measurement_points) * 100  # Updating of the process bar.
            progressbar.after(1, progress(currentValue))
            progressbar.update()

    if type_of_the_analysis == 2 or type_of_the_analysis == 3:
        my_window2.destroy()
    #----------------------------------------------------------------------------------------------------------------------

    cube = []  # I place the maximum dB values of each measurment point in one array. This values are gonna be displayed on the colormap.
    for k in range(number_of_measurement_points):
        cube.append(dB[k].max())

    array = np.array(cube).reshape(y,x)  # With this line I reshape the one row cube array which contains the max. dB values to an array which contains x number of rows and y number of columns.

    for k in range(0,y):  # In order to display the measurement points according to the order how they come from the spectrum analyzer, every second row has to be flipped.
        if k % 2 != 0:  # With this the snake shaped order of the measuring points is obtained.
            array[k] = np.flip(array[k])

    def onclick(event):
        p = int(round((event.xdata + (step_mm / 2)) * (x / (colormap_A + step_mm)) + 0.5))  # This part is to plot the frequency-dB diagrams when clicking on a specific point of the colormap.
        q = int(round((event.ydata + (step_mm / 2)) * (y / (colormap_B + step_mm)) + 0.5))

        if q % 2 != 0:
            q = -1 + (q - 1) * x
            r = q + p  # r is the value which is filled in the colormap. Actually this indicates the number of measurement point which belongs to the segment of the colormap where you click.
        else:
            q = q * x
            r = q - p

        deleted_frequencies = np.array([])  # to determine the position of the deleted frequencies. In these points a red dot is displayed on the x-axis
        for k in range(np.size(indices_to_be_deleted)):
            deleted_frequencies = np.append(deleted_frequencies, frequency[indices_to_be_deleted[k][0]])

        y_axis_of_red_dots = np.array([])   # the y-axis of the red dots has to be at the zero origin. The origin of the plot is the smallest value of the measured datas minus 5.
        for k in range(np.size(indices_to_be_deleted)): # this value is placed in an array as many times as many indices are deleted.
            y_axis_of_red_dots = np.append(y_axis_of_red_dots, np.amin(dB[5])-5)

        plt.figure(2)
        plt.ylim((np.amin(dB[5]) - 5), 0)   # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
        plt.plot(frequency, dB[r])  # The frequency axis is always the same but the dB values are changing in terms of where the clicking is on the colormap. This r values gives the number of measurement point from where the dB values have to be called.
        plt.xlabel('Frequency [Hz]')  # These are the labels and title of the plot which comes after clicking on a point of the colormap.
        plt.ylabel('Amplitude [dB]')
        plt.title('EMC Measurement')
        plt.scatter(deleted_frequencies, y_axis_of_red_dots, marker='o', s=10., c='r')  # This is to display the red dots on the edited plot.
        plt.show()

    fig = plt.figure(1)  # This is the first figure after analysing the data. This calls the colormap.
    ax = fig.add_subplot()
    alpha = 1  # If there is no background picture, no need to bedim the colormap. In that case the colormap will not be transparent ==> (alpha = 1).
    if there_is_a_picture is True:
        board = Image.open(picture_of_the_PCB)
        ax.imshow(board, alpha=1, extent=[0, A, B, 0])
        alpha = 0.7  # if there is background picture, the colormap is bedimmed a bit
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')

    colormap_A = ((round_half_up(A / step_mm)) * step_mm)
    colormap_B = ((round_half_up(B / step_mm)) * step_mm)

    plt.imshow(array, cmap='YlOrRd', alpha=alpha, extent=[-step_mm / 2, colormap_A + step_mm / 2, colormap_B + step_mm / 2, -step_mm / 2])  # The maximum dB values are saved in the "array" variable.
    clb = plt.colorbar(mappable=None, cax=None, ax=None)
    clb.set_label('Amplitude [dB]', labelpad=0, y=0.5, rotation=90, )
    plt.connect('button_press_event', onclick)  # Here I connect the onclick function to this colormap.

    plt.xlabel("A side of the PCB [mm]")
    plt.ylabel("B side of the PCB [mm]")

    plt.tight_layout()
    wm = plt.get_current_fig_manager()  # To display the figure in maximum size.
    wm.window.state('zoomed')

    plt.show()
    #-------------------------------------------------------------------------------------------------

def Continue_without_changes(): # This loop gets executed when clicking on the continue button if the user chose to not change anything on the consideration of ambient radiation or on the analyzation type
    plt.close('all')
    input_values_path = str(e1.get())
    displayed_data = str(e2.get())
    indices_to_be_deleted_path = str(e4.get())
    picture_of_the_PCB = str(e5.get())

    input_values = openpyxl.load_workbook(input_values_path)     # Importing the input values.
    sheet = input_values.active
    A = float(sheet.cell(4, 2).value)
    B = float(sheet.cell(5, 2).value)
    start_freq_level = float(sheet.cell(11, 2).value)
    stop_freq_level = float(sheet.cell(12, 2).value)
    step_mm = float(sheet.cell(14, 2).value)
    Quit()

    x = int(round_half_up(A / step_mm) + 1)  # This x value gives the number of cubes on the color map on the horizontal axis
    y = int(round_half_up(B / step_mm) + 1)  # This y value gives the number of cubes on the color map on the vertical axis

    number_of_measurement_points = int(x * y)  # When the robot stops (after each step) python reads the values from the spectrum analyzer. This gives the number of measurements.

    indices_to_be_deleted = np.load(indices_to_be_deleted_path) # Loading of the numpy array which contains the number incices to be deleted.

    there_is_a_picture = False

    if os.path.exists(picture_of_the_PCB) is True:
        there_is_a_picture = True  # if you input the location of a picture it means there will be a background behind the colormap. there_is_a_picture boolean is used to distinguish between displaying a background or no.

    frequency = np.linspace(float(start_freq_level), float(stop_freq_level), int(number_of_points_on_the_analyzer))  # Processing the values from the excel sheet into arrays for the easier analysation.

    dB = np.load(displayed_data)

    cube = []  # I place the maximum dB values of each measurment point in one array. This values are gonna be displayed on the colormap.
    for k in range(number_of_measurement_points):
        cube.append(dB[k].max())

    array = np.array(cube).reshape(y,x)  # With this line I reshape the one row cube array which contains the max. dB values to an array which contains x number of rows and y number of columns.

    for k in range(0,y):  # In order to display the measurement points according to the order how they come from the spectrum analyzer, every second row has to be flipped.
        if k % 2 != 0:  # With this the snake shaped order of the measuring points is obtained.
            array[k] = np.flip(array[k])

    def onclick(event):
        p = int(round((event.xdata + (step_mm / 2)) * (x / (colormap_A + step_mm)) + 0.5))  # This part is to plot the frequency-dB diagrams when clicking on a specific point of the colormap.
        q = int(round((event.ydata + (step_mm / 2)) * (y / (colormap_B + step_mm)) + 0.5))

        if q % 2 != 0:
            q = -1 + (q - 1) * x
            r = q + p  # r is the value which is filled in the colormap. Actually this indicates the number of measurement point which belongs to the segment of the colormap where you click.
        else:
            q = q * x
            r = q - p

        deleted_frequencies = np.array([])  # to dtermine the position of the deleted frequencies. In these points a red dot is displayed on the x-axis
        for k in range(np.size(indices_to_be_deleted)):
            deleted_frequencies = np.append(deleted_frequencies, frequency[indices_to_be_deleted[k][0]])

        y_axis_of_red_dots = np.array([])   # the y-axis of the red dots has to be at the zero origin. The origin of the plot is the smallest value of the measured datas minus 5.
        for k in range(np.size(indices_to_be_deleted)): # this value is placed in an array as many times as many indices are deleted.
            y_axis_of_red_dots = np.append(y_axis_of_red_dots, np.amin(dB[5])-5)

        plt.figure(2)
        plt.ylim((np.amin(dB[5]) - 5), 0)   # To set the y axis to be between 0 and the least value of the amplidudes. It makes is to be easier visualizable.
        plt.plot(frequency, dB[r])  # The frequency axis is always the same but the dB values are changing in terms of where the clicking is on the colormap. This r values gives the number of measurement point from where the dB values have to be called.
        plt.xlabel('Frequency [Hz]')  # These are the labels and title of the plot which comes after clicking on a point of the colormap.
        plt.ylabel('Amplitude [dB]')
        plt.title('EMC Measurement')
        plt.scatter(deleted_frequencies, y_axis_of_red_dots, marker='o', s=10., c='r')  # This is to display the red dots on the edited plot.
        plt.show()

    fig = plt.figure(1)  # This is the first figure after analysing the data. This calls the colormap.
    ax = fig.add_subplot()
    alpha = 1  # If there is no background picture, no need to bedim the colormap. In that case the colormap will not be transparent ==> (alpha = 1).
    if there_is_a_picture is True:
        board = Image.open(picture_of_the_PCB)
        ax.imshow(board, alpha=1, extent=[0, A, B, 0])
        alpha = 0.7  # if there is background picture, the colormap is bedimmed a bit
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')

    colormap_A = ((round_half_up(A / step_mm)) * step_mm)
    colormap_B = ((round_half_up(B / step_mm)) * step_mm)

    plt.imshow(array, cmap='YlOrRd', alpha=alpha, extent=[-step_mm / 2, colormap_A + step_mm / 2, colormap_B + step_mm / 2, -step_mm / 2])  # The maximum dB values are saved in the "array" variable.
    clb = plt.colorbar(mappable=None, cax=None, ax=None)
    clb.set_label('Amplitude [dB]', labelpad=0, y=0.5, rotation=90, )
    plt.connect('button_press_event', onclick)  # Here I connect the onclick function to this colormap.

    plt.xlabel("A side of the PCB [mm]")
    plt.ylabel("B side of the PCB [mm]")

    plt.tight_layout()
    wm = plt.get_current_fig_manager()  # To display the figure in maximum size.
    wm.window.state('zoomed')

    plt.show()

def Quit():
    my_window.destroy()

# Here the displayed window is being designed.
#-----------------------------------------------------------------------------------------------------------------------------
my_window = tk.Tk()
my_window.title("Evaluation of measurements")
my_window.state('zoomed')
tesat = ImageTk.PhotoImage(Image.open("tesat.jpg"))
tk.Label(my_window, image=tesat).grid(row=1, column=6) # Tesat image on the right side.
response = tkinter.messagebox.askquestion('Recalling of previously executed measurement', 'Would you like to make changes on the consideration of the ambient radiation or the analyses type?')
if 'yes' in response:
    process = "yes"
    tk.Label(my_window, text="1. Path of the input values: ", font=2).grid(row=2, column=1)
    tk.Label(my_window, text="2. Path of the measurement data: ", font=2).grid(row=3, column=1)
    tk.Label(my_window, text="3. Path of the ambient radiation: ", font=2).grid(row=4, column=1)
    tk.Label(my_window, text="4. Path of the indices to be deleted: ", font=2).grid(row=5, column=1)
    tk.Label(my_window, text="5. Picture of the PCB: ", font=2).grid(row=6, column=1)
    tk.Label(my_window, text="6. Selection of the ambient radiation frequency ranges: ", font=2).grid(row=7, column=1)
    tk.Label(my_window, text="7. Type of the analysis: ", font=2).grid(row=8, column=1)
    tk.Label(my_window, text="", font=2).grid(row=9, column=1)

    e1 = tk.Entry(my_window)    # Setting of the input fields.
    e2 = tk.Entry(my_window)
    e3 = tk.Entry(my_window)
    e4 = tk.Entry(my_window)
    e5 = tk.Entry(my_window)
    e6 = tk.Entry(my_window)
    e7 = IntVar(my_window)    # Checkbox of the 16th question. It can be either Single, Maximum or Average

    e1.grid(row=2, column=2, pady=3)
    e2.grid(row=3, column=2, pady=3)
    e3.grid(row=4, column=2, pady=3)
    e4.grid(row=5, column=2, pady=3)
    e5.grid(row=6, column=2, pady=3)
    e6.grid(row=7, column=2, pady=3)

if 'no' in response:
    process = "no"
    tk.Label(my_window, text="1. Path of the input values: ", font=2).grid(row=2, column=1)
    tk.Label(my_window, text="2. Path of the displayed data: ", font=2).grid(row=3, column=1)
    tk.Label(my_window, text="3. Path of the indices to be deleted: ", font=2).grid(row=4, column=1)
    tk.Label(my_window, text="4. Picture of the PCB: ", font=2).grid(row=5, column=1)
    tk.Label(my_window, text="", font=2).grid(row=6, column=1)

    e1 = tk.Entry(my_window)    # Setting of the input fields.
    e2 = tk.Entry(my_window)
    e4 = tk.Entry(my_window)
    e5 = tk.Entry(my_window)

    e1.grid(row=2, column=2, pady=3)
    e2.grid(row=3, column=2, pady=3)
    e4.grid(row=4, column=2, pady=3)
    e5.grid(row=5, column=2, pady=3)

# Settings of the buttons. All the buttons call a separate definition depending on the function it stands for.
if process == "yes":  # functions used if the user would like to change on the consideration of ambient radiation or on the analyzes type.
    tk.Button(my_window, text='Print path', font=2, command=print_path_input_values, width=8).grid(row=2, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_measurement_data, width=8).grid(row=3, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_ambient_radiation, width=8).grid(row=4, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_indices_to_be_deleted, width=8).grid(row=5, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_picture, width=8).grid(row=6, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Start', font=2, command=selection_of_ambient_radiation, width=8).grid(row=7, column=3, padx=3, pady=3)
    Radiobutton(my_window, text="Single", font=2, variable=e7, value=1).grid(row=8, column=2, padx=3, pady=3)
    Radiobutton(my_window, text="Maximum", font=2, variable=e7, value=2).grid(row=8, column=3, padx=3, pady=3)
    Radiobutton(my_window, text="Average", font=2, variable=e7, value=3).grid(row=8, column=4, padx=3, pady=3)
    tk.Button(my_window, text='Continue', font=2, command=Continue_with_changes, width=8).grid(row=10, column=5, padx=3, pady=3)
    tk.Button(my_window, text='Quit', font=2, command=Quit, width=8).grid(row=10, column=4, padx=3, pady=3)

    my_window.grid_rowconfigure(0, weight=1)  # arranging the the content of the window in the center
    my_window.grid_rowconfigure(11, weight=1)
    my_window.grid_columnconfigure(0, weight=1)
    my_window.grid_columnconfigure(7, weight=1)

if process == "no": # functions used if the user wants to display the result of the measurement without any change.
    tk.Button(my_window, text='Print path', font=2, command=print_path_input_values, width=8).grid(row=2, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_displayed_data, width=8).grid(row=3, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_indices_to_be_deleted, width=8).grid(row=4, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Print path', font=2, command=print_path_picture, width=8).grid(row=5, column=3, padx=3, pady=3)
    tk.Button(my_window, text='Continue', font=2, command=Continue_without_changes, width=8).grid(row=6, column=5, padx=3, pady=3)
    tk.Button(my_window, text='Quit', font=2,command=Quit, width=8).grid(row=6, column=4, padx=3, pady=3)

    my_window.grid_rowconfigure(0, weight=1)  # arranging the the content of the window in the center
    my_window.grid_rowconfigure(8, weight=1)
    my_window.grid_columnconfigure(0, weight=1)
    my_window.grid_columnconfigure(7, weight=1)

my_window.mainloop()